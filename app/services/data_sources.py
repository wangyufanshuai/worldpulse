from __future__ import annotations

import io
import json
import zipfile
from datetime import date, timedelta
from pathlib import Path
from time import time

import numpy as np
import pandas as pd
import requests

from app.services.config import load_config

NASA_GISTEMP_URL = "https://data.giss.nasa.gov/gistemp/tabledata_v4/GLB.Ts+dSST.csv"
NOAA_CO2_URL = "https://gml.noaa.gov/webdata/ccgg/trends/co2/co2_mm_mlo.csv"
NOAA_ONI_URL = "https://www.cpc.ncep.noaa.gov/data/indices/oni.ascii.txt"
WORLD_BANK_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
)
GDELT_DOC_URL = "https://api.gdeltproject.org/api/v2/doc/doc"
UCDP_GED_URL = "https://ucdp.uu.se/downloads/ged/ged251-csv.zip"
WORLD_UNCERTAINTY_INDEX_URL = "https://worlduncertaintyindex.com/wp-content/uploads/2026/05/WUI_M_dataset_2026_04.xlsx"
FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv"
CACHE_DIR = Path("data/cache")


def load_world_data(days: int = 420) -> tuple[pd.DataFrame, dict[str, str]]:
    end = date.today()
    start = end - timedelta(days=days)
    dates = pd.bdate_range(start=start, end=end)
    frame = pd.DataFrame({"date": dates})
    sources: dict[str, str] = {}

    config = load_config()
    for item in config["financial"]["series"]:
        series, source = _load_financial_series(item, dates)
        frame[item["key"]] = series
        sources[item["key"]] = source

    for family in ["climate", "geopolitical", "ecology", "macro"]:
        for item in config[family]["synthetic"]:
            series, source = _load_real_or_demo_signal(item["key"], dates)
            frame[item["key"]] = series
            sources[item["key"]] = source

    return frame.sort_values("date").reset_index(drop=True), sources


def _load_real_or_demo_signal(key: str, dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    loaders = {
        "temp_anomaly": _load_nasa_temperature_signal,
        "ocean_heat": _load_nasa_ocean_heat_proxy,
        "co2_pressure": _load_noaa_co2_signal,
        "enso_stress": _load_noaa_enso_signal,
        "conflict_intensity": lambda idx: _load_gdelt_signal(
            "conflict_intensity",
            "(war OR conflict OR attack OR missile OR protest OR coup)",
            idx,
        ),
        "policy_uncertainty": _load_world_uncertainty_signal,
        "food_pressure": _load_world_bank_food_signal,
        "drought_stress": _load_world_bank_drought_proxy,
        "fertilizer_pressure": _load_world_bank_fertilizer_signal,
        "credit_spread": lambda idx: _load_fred_scaled_signal("BAMLH0A0HYM2", idx, low=3.0, high=9.0),
        "yield_curve": _load_yield_curve_signal,
        "dollar_stress": _load_dollar_stress_signal,
        "gas_pressure": lambda idx: _load_fred_index_signal("DHHNGSP", idx, source_name="FRED:DHHNGSP natural gas"),
    }
    loader = loaders.get(key)
    if loader:
        try:
            series, source = loader(dates)
            if series.notna().sum() > 20:
                return series.clip(0, 1).ffill().bfill(), source
            if key == "conflict_intensity":
                return _load_ucdp_conflict_signal(dates)
        except Exception:
            if key == "conflict_intensity":
                try:
                    return _load_ucdp_conflict_signal(dates)
                except Exception:
                    pass
    return _synthetic_signal(key, dates), "deterministic-demo"


def _load_financial_series(item: dict, dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    fred_id = item.get("fred_id")
    if fred_id:
        try:
            series = _load_fred_series(fred_id, dates)
            if series.notna().sum() > 30:
                return series, f"FRED:{fred_id}"
        except Exception:
            pass

    world_bank_code = item.get("world_bank_price_code")
    if world_bank_code:
        try:
            series = _load_world_bank_price_series(world_bank_code, dates)
            if series.notna().sum() > 30:
                return series, f"World Bank Pink Sheet:{world_bank_code}"
        except Exception:
            pass

    symbol = item["symbol"]
    try:
        import yfinance as yf

        data = yf.download(
            symbol,
            start=dates.min().date().isoformat(),
            end=(dates.max().date() + timedelta(days=1)).isoformat(),
            progress=False,
            auto_adjust=True,
            threads=False,
        )
        if not data.empty:
            if isinstance(data.columns, pd.MultiIndex):
                data.columns = [col[0] for col in data.columns]
            close = data["Close"].reset_index()
            close.columns = ["date", "close"]
            close["date"] = pd.to_datetime(close["date"]).dt.tz_localize(None)
            merged = pd.DataFrame({"date": dates}).merge(close, on="date", how="left")
            series = merged["close"].ffill().bfill()
            if series.notna().sum() > 30:
                return series, f"yfinance:{symbol}"
    except Exception:
        pass

    return _synthetic_price(symbol, dates), "deterministic-demo"


def _load_fred_series(series_id: str, dates: pd.DatetimeIndex) -> pd.Series:
    filename = f"fred_{series_id}.csv"
    url = f"{FRED_CSV_URL}?id={series_id}"
    raw = _read_text_cache(filename, url, ttl_seconds=6 * 3600)
    frame = pd.read_csv(io.StringIO(raw))
    if frame.empty or series_id not in frame.columns:
        raise ValueError(f"FRED series {series_id} returned no usable data")
    frame = frame.rename(columns={"observation_date": "date", series_id: "close"})
    frame["date"] = pd.to_datetime(frame["date"])
    frame["close"] = pd.to_numeric(frame["close"].replace(".", np.nan), errors="coerce")
    frame = frame.dropna(subset=["close"]).sort_values("date")
    return _align_price_frame(frame[["date", "close"]], dates)


def _load_world_bank_price_series(code: str, dates: pd.DatetimeIndex) -> pd.Series:
    prices = _load_world_bank_prices()
    if code not in prices.columns:
        raise ValueError(f"World Bank price code not found: {code}")
    frame = prices[["date", code]].rename(columns={code: "close"})
    frame["close"] = pd.to_numeric(frame["close"], errors="coerce")
    frame = frame.dropna(subset=["close"])
    return _align_price_frame(frame, dates)


def _load_nasa_temperature_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    raw = _read_text_cache("nasa_gistemp.csv", NASA_GISTEMP_URL, ttl_seconds=24 * 3600)
    monthly = _parse_gistemp_csv(raw)
    monthly["risk"] = _scale_series(monthly["value"], low=0.3, high=1.6)
    return _align_monthly(monthly[["date", "risk"]], dates), "NASA GISTEMP"


def _load_nasa_ocean_heat_proxy(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    raw = _read_text_cache("nasa_gistemp.csv", NASA_GISTEMP_URL, ttl_seconds=24 * 3600)
    monthly = _parse_gistemp_csv(raw)
    smoothed = monthly["value"].rolling(12, min_periods=3).mean()
    acceleration = smoothed.diff(12).fillna(0)
    monthly["risk"] = (_scale_series(smoothed, low=0.3, high=1.6) * 0.75 + _scale_series(acceleration, low=-0.15, high=0.25) * 0.25)
    return _align_monthly(monthly[["date", "risk"]], dates), "NASA GISTEMP-derived ocean heat proxy"


def _load_noaa_co2_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    raw = _read_text_cache("noaa_co2_mm_mlo.csv", NOAA_CO2_URL, ttl_seconds=7 * 24 * 3600)
    frame = pd.read_csv(io.StringIO(raw), comment="#")
    frame["date"] = pd.to_datetime(dict(year=frame["year"], month=frame["month"], day=1))
    frame["average"] = pd.to_numeric(frame["average"], errors="coerce").replace(-99.99, np.nan)
    frame = frame.dropna(subset=["average"]).sort_values("date")
    momentum = frame["average"].diff(12).fillna(0)
    frame["risk"] = (_scale_series(frame["average"], low=390, high=435) * 0.7 + _scale_series(momentum, low=1.5, high=4.0) * 0.3)
    return _align_monthly(frame[["date", "risk"]], dates), "NOAA Mauna Loa CO2"


def _load_noaa_enso_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    raw = _read_text_cache("noaa_oni.txt", NOAA_ONI_URL, ttl_seconds=7 * 24 * 3600)
    frame = pd.read_csv(io.StringIO(raw), sep=r"\s+")
    season_month = {
        "DJF": 1,
        "JFM": 2,
        "FMA": 3,
        "MAM": 4,
        "AMJ": 5,
        "MJJ": 6,
        "JJA": 7,
        "JAS": 8,
        "ASO": 9,
        "SON": 10,
        "OND": 11,
        "NDJ": 12,
    }
    frame["month"] = frame["SEAS"].map(season_month)
    frame["date"] = pd.to_datetime(dict(year=frame["YR"], month=frame["month"], day=1))
    frame["ANOM"] = pd.to_numeric(frame["ANOM"], errors="coerce")
    frame = frame.dropna(subset=["ANOM"]).sort_values("date")
    frame["risk"] = _scale_series(frame["ANOM"].abs(), low=0.4, high=2.2)
    return _align_monthly(frame[["date", "risk"]], dates), "NOAA CPC Oceanic Nino Index"


def _parse_gistemp_csv(raw: str) -> pd.DataFrame:
    data = pd.read_csv(io.StringIO(raw), skiprows=1)
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    rows = []
    for _, row in data.iterrows():
        year = int(row["Year"])
        for month_number, month in enumerate(months, 1):
            value = pd.to_numeric(row[month], errors="coerce")
            if pd.notna(value):
                rows.append({"date": pd.Timestamp(year=year, month=month_number, day=1), "value": float(value)})
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def _load_world_bank_food_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    indices = _load_world_bank_indices()
    indices["risk"] = _risk_from_index(indices["iFOOD"])
    return _align_monthly(indices[["date", "risk"]], dates), "World Bank Pink Sheet iFOOD"


def _load_world_bank_drought_proxy(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    indices = _load_world_bank_indices()
    blend = indices[["iGRAINS", "iFATS_OILS", "iOTHERFOOD"]].mean(axis=1)
    indices["risk"] = _risk_from_index(blend)
    return _align_monthly(indices[["date", "risk"]], dates), "World Bank Pink Sheet agriculture proxy"


def _load_world_bank_fertilizer_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    indices = _load_world_bank_indices()
    indices["risk"] = _risk_from_index(indices["iFERTILIZERS"])
    return _align_monthly(indices[["date", "risk"]], dates), "World Bank Pink Sheet iFERTILIZERS"


def _load_fred_scaled_signal(series_id: str, dates: pd.DatetimeIndex, low: float, high: float) -> tuple[pd.Series, str]:
    filename = f"fred_{series_id}.csv"
    url = f"{FRED_CSV_URL}?id={series_id}"
    raw = _read_text_cache(filename, url, ttl_seconds=6 * 3600)
    frame = pd.read_csv(io.StringIO(raw))
    frame = frame.rename(columns={"observation_date": "date", series_id: "value"})
    frame["date"] = pd.to_datetime(frame["date"])
    frame["value"] = pd.to_numeric(frame["value"].replace(".", np.nan), errors="coerce")
    frame = frame.dropna(subset=["value"]).sort_values("date")
    frame["risk"] = _scale_series(frame["value"], low=low, high=high)
    return _align_daily(frame[["date", "risk"]], dates), f"FRED:{series_id}"


def _load_fred_index_signal(series_id: str, dates: pd.DatetimeIndex, source_name: str | None = None) -> tuple[pd.Series, str]:
    filename = f"fred_{series_id}.csv"
    url = f"{FRED_CSV_URL}?id={series_id}"
    raw = _read_text_cache(filename, url, ttl_seconds=6 * 3600)
    frame = pd.read_csv(io.StringIO(raw))
    frame = frame.rename(columns={"observation_date": "date", series_id: "value"})
    frame["date"] = pd.to_datetime(frame["date"])
    frame["value"] = pd.to_numeric(frame["value"].replace(".", np.nan), errors="coerce")
    frame = frame.dropna(subset=["value"]).sort_values("date")
    frame["risk"] = _risk_from_index(frame["value"])
    return _align_daily(frame[["date", "risk"]], dates), source_name or f"FRED:{series_id}"


def _load_yield_curve_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    filename = "fred_T10Y2Y.csv"
    url = f"{FRED_CSV_URL}?id=T10Y2Y"
    raw = _read_text_cache(filename, url, ttl_seconds=6 * 3600)
    frame = pd.read_csv(io.StringIO(raw))
    frame = frame.rename(columns={"observation_date": "date", "T10Y2Y": "value"})
    frame["date"] = pd.to_datetime(frame["date"])
    frame["value"] = pd.to_numeric(frame["value"].replace(".", np.nan), errors="coerce")
    frame = frame.dropna(subset=["value"]).sort_values("date")
    frame["risk"] = _scale_series(-frame["value"], low=-1.5, high=1.0)
    return _align_daily(frame[["date", "risk"]], dates), "FRED:T10Y2Y"


def _load_dollar_stress_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    return _load_fred_index_signal("DTWEXBGS", dates, source_name="FRED:DTWEXBGS broad dollar")


def _load_world_bank_indices() -> pd.DataFrame:
    content = _read_bytes_cache("world_bank_pink_sheet.xlsx", WORLD_BANK_PINK_SHEET_URL, ttl_seconds=7 * 24 * 3600)
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Monthly Indices"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=10, max_row=10))]
    rows = []
    for row in sheet.iter_rows(min_row=11, values_only=True):
        if not row or not row[0]:
            continue
        item = dict(zip(headers, row))
        date_text = str(row[0])
        if "M" not in date_text:
            continue
        year_text, month_text = date_text.split("M", 1)
        item["date"] = pd.Timestamp(year=int(year_text), month=int(month_text), day=1)
        rows.append(item)
    frame = pd.DataFrame(rows)
    numeric_cols = [col for col in ["iFOOD", "iGRAINS", "iFATS_OILS", "iOTHERFOOD"] if col in frame.columns]
    for col in numeric_cols:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)


def _load_world_bank_prices() -> pd.DataFrame:
    content = _read_bytes_cache("world_bank_pink_sheet.xlsx", WORLD_BANK_PINK_SHEET_URL, ttl_seconds=7 * 24 * 3600)
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Monthly Prices"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=7, max_row=7))]
    rows = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or not row[0]:
            continue
        item = dict(zip(headers, row))
        date_text = str(row[0])
        if "M" not in date_text:
            continue
        year_text, month_text = date_text.split("M", 1)
        item["date"] = pd.Timestamp(year=int(year_text), month=int(month_text), day=1)
        rows.append(item)
    frame = pd.DataFrame(rows)
    for col in frame.columns:
        if col != "date":
            frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)


def _risk_from_index(series: pd.Series) -> pd.Series:
    momentum = series.pct_change(6).fillna(0)
    percentile = series.rolling(120, min_periods=24).rank(pct=True).fillna(0.5)
    return (percentile * 0.7 + _scale_series(momentum, low=-0.15, high=0.35) * 0.3).clip(0, 1)


def _load_gdelt_signal(cache_key: str, query: str, dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    params = {
        "query": query,
        "mode": "timelinevolraw",
        "format": "json",
        "timespan": "18months",
    }
    payload = _read_json_cache(f"gdelt_{cache_key}.json", GDELT_DOC_URL, params=params, ttl_seconds=6 * 3600)
    points = _parse_gdelt_timeline(payload)
    if points.empty:
        raise ValueError("GDELT timeline returned no parseable points")
    points["risk"] = _risk_from_index(points["value"])
    daily = _align_daily(points[["date", "risk"]], dates)
    return daily, f"GDELT DOC timeline: {cache_key}"


def _load_ucdp_conflict_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    monthly_path = _cache_path("ucdp_ged_monthly.csv")
    if _cache_fresh(monthly_path, ttl_seconds=7 * 24 * 3600):
        monthly = pd.read_csv(monthly_path)
        monthly["date"] = pd.to_datetime(monthly["date"])
    else:
        content = _read_bytes_cache("ucdp_ged251_csv.zip", UCDP_GED_URL, ttl_seconds=30 * 24 * 3600)
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            csv_name = next(name for name in archive.namelist() if name.lower().endswith(".csv"))
            with archive.open(csv_name) as file:
                events = pd.read_csv(
                    file,
                    usecols=["date_start", "best", "deaths_civilians", "region"],
                    parse_dates=["date_start"],
                )
        events["date"] = events["date_start"].dt.to_period("M").dt.to_timestamp()
        events["best"] = pd.to_numeric(events["best"], errors="coerce").fillna(0)
        events["deaths_civilians"] = pd.to_numeric(events["deaths_civilians"], errors="coerce").fillna(0)
        monthly = (
            events.groupby("date")
            .agg(events=("best", "size"), deaths=("best", "sum"), civilian_deaths=("deaths_civilians", "sum"))
            .reset_index()
        )
        monthly.to_csv(monthly_path, index=False)

    intensity = np.log1p(monthly["events"]) * 0.35 + np.log1p(monthly["deaths"]) * 0.45 + np.log1p(monthly["civilian_deaths"]) * 0.20
    monthly["risk"] = _risk_from_index(intensity)
    return _align_monthly(monthly[["date", "risk"]], dates), "UCDP GED 25.1 conflict events"


def _load_world_uncertainty_signal(dates: pd.DatetimeIndex) -> tuple[pd.Series, str]:
    content = _read_bytes_cache("world_uncertainty_index.xlsx", WORLD_UNCERTAINTY_INDEX_URL, ttl_seconds=30 * 24 * 3600)
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["F1"]
    rows = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        rows.append({"date": pd.to_datetime(row[0]), "value": float(row[1])})
    frame = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
    frame["risk"] = _risk_from_index(frame["value"])
    return _align_monthly(frame[["date", "risk"]], dates), "World Uncertainty Index"


def _parse_gdelt_timeline(payload: dict) -> pd.DataFrame:
    candidates = []
    for timeline in payload.get("timeline", []):
        for point in timeline.get("data", []):
            candidates.append(point)
    if not candidates and "data" in payload:
        candidates = payload["data"]
    rows = []
    for point in candidates:
        date_value = point.get("date") or point.get("datetime") or point.get("timestamp")
        value = point.get("value") or point.get("count") or point.get("Volume Intensity")
        if date_value is None or value is None:
            continue
        parsed_date = _parse_gdelt_date(str(date_value))
        if parsed_date is None:
            continue
        rows.append({"date": parsed_date, "value": float(value)})
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True) if rows else pd.DataFrame()


def _parse_gdelt_date(value: str) -> pd.Timestamp | None:
    for fmt in ["%Y%m%d%H%M%S", "%Y%m%d%H%M", "%Y%m%d", "%Y-%m-%d"]:
        try:
            return pd.to_datetime(value, format=fmt)
        except Exception:
            continue
    try:
        return pd.to_datetime(value)
    except Exception:
        return None


def _seed(key: str) -> int:
    return abs(hash(key)) % (2**32)


def _synthetic_price(key: str, dates: pd.DatetimeIndex) -> pd.Series:
    rng = np.random.default_rng(_seed(key))
    drift = 0.00015
    vol = 0.012
    if key in {"^VIX"}:
        base = 18
        noise = rng.normal(0, 1.2, len(dates))
        cycle = 5 * np.sin(np.linspace(0, 8 * np.pi, len(dates)))
        spikes = rng.choice([0, 0, 0, 8, 15], len(dates), p=[0.9, 0.04, 0.03, 0.02, 0.01])
        return pd.Series(np.maximum(10, base + cycle + noise + spikes))
    shocks = rng.normal(drift, vol, len(dates))
    return pd.Series(100 * np.exp(np.cumsum(shocks)))


def _synthetic_signal(key: str, dates: pd.DatetimeIndex) -> pd.Series:
    rng = np.random.default_rng(_seed(key))
    n = len(dates)
    base_level = {
        "temp_anomaly": 0.55,
        "ocean_heat": 0.62,
        "co2_pressure": 0.58,
        "enso_stress": 0.35,
        "conflict_intensity": 0.48,
        "policy_uncertainty": 0.45,
        "drought_stress": 0.42,
        "food_pressure": 0.46,
        "fertilizer_pressure": 0.44,
        "credit_spread": 0.38,
        "yield_curve": 0.52,
        "dollar_stress": 0.45,
        "gas_pressure": 0.40,
    }.get(key, 0.5)
    cycle = 0.16 * np.sin(np.linspace(0, 5 * np.pi, n) + (_seed(key) % 11))
    trend = np.linspace(-0.04, 0.08, n)
    noise = rng.normal(0, 0.035, n)
    values = np.clip(base_level + cycle + trend + noise, 0, 1)
    return pd.Series(values)


def _align_monthly(monthly: pd.DataFrame, dates: pd.DatetimeIndex) -> pd.Series:
    frame = monthly.copy()
    frame["date"] = pd.to_datetime(frame["date"])
    frame = frame.sort_values("date").set_index("date")
    daily = frame.reindex(pd.date_range(frame.index.min(), dates.max(), freq="D")).ffill()
    return daily.reindex(dates, method="ffill")["risk"].ffill().bfill().reset_index(drop=True)


def _align_daily(daily_frame: pd.DataFrame, dates: pd.DatetimeIndex) -> pd.Series:
    frame = daily_frame.copy()
    frame["date"] = pd.to_datetime(frame["date"]).dt.tz_localize(None)
    frame = frame.sort_values("date").drop_duplicates("date").set_index("date")
    daily = frame.reindex(pd.date_range(min(frame.index.min(), dates.min()), dates.max(), freq="D")).ffill()
    return daily.reindex(dates, method="ffill")["risk"].ffill().bfill().reset_index(drop=True)


def _align_price_frame(price_frame: pd.DataFrame, dates: pd.DatetimeIndex) -> pd.Series:
    frame = price_frame.copy()
    frame["date"] = pd.to_datetime(frame["date"]).dt.tz_localize(None)
    frame = frame.sort_values("date").drop_duplicates("date").set_index("date")
    daily = frame.reindex(pd.date_range(min(frame.index.min(), dates.min()), dates.max(), freq="D")).ffill()
    return daily.reindex(dates, method="ffill")["close"].ffill().bfill().reset_index(drop=True)


def _scale_series(series: pd.Series, low: float, high: float) -> pd.Series:
    return ((series - low) / (high - low)).clip(0, 1).fillna(0.5)


def _read_text_cache(filename: str, url: str, ttl_seconds: int) -> str:
    path = _cache_path(filename)
    if _cache_fresh(path, ttl_seconds):
        return path.read_text(encoding="utf-8")
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    text = response.text
    path.write_text(text, encoding="utf-8")
    return text


def _read_bytes_cache(filename: str, url: str, ttl_seconds: int) -> bytes:
    path = _cache_path(filename)
    if _cache_fresh(path, ttl_seconds):
        return path.read_bytes()
    response = requests.get(url, timeout=45)
    response.raise_for_status()
    content = response.content
    path.write_bytes(content)
    return content


def _read_json_cache(filename: str, url: str, params: dict, ttl_seconds: int) -> dict:
    path = _cache_path(filename)
    if _cache_fresh(path, ttl_seconds):
        return json.loads(path.read_text(encoding="utf-8"))
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    payload = response.json()
    path.write_text(json.dumps(payload), encoding="utf-8")
    return payload


def _cache_path(filename: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return CACHE_DIR / filename


def _cache_fresh(path: Path, ttl_seconds: int) -> bool:
    return path.exists() and time() - path.stat().st_mtime < ttl_seconds
